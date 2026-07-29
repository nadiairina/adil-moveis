import openpyxl

file_path = '../Excel de produtos - Site-finalporagora.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

print("Extracting from", file_path)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        colors = []
        for cell in row:
            if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                color_type = cell.fill.fgColor.type
                if color_type == 'rgb':
                    colors.append(cell.fill.fgColor.rgb)
                elif color_type == 'theme':
                    colors.append(f"theme-{cell.fill.fgColor.theme}")
        
        # Filter out clear/white background
        valid_colors = [c for c in colors if c not in ('00000000', 'FFFFFFFF')]
        if valid_colors:
            row_values = [str(c.value) if c.value is not None else '' for c in row]
            if any(row_values):
                print(f"Row {row[0].row} [Colors: {set(valid_colors)}]: {row_values}")

