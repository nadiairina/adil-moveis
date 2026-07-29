import openpyxl

wb = openpyxl.load_workbook('../Excel de produtos - Site-finalporagora.xlsx', data_only=True)
total_highlighted = 0

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    count = 0
    highlighted_items = []
    
    for row in sheet.iter_rows():
        is_highlighted = False
        item_name = ""
        for cell in row:
            # We assume column C (index 2) usually has the name, but let's just grab the first string we see if highlighted
            if type(cell.value) == str and len(cell.value) > 2:
                if not item_name:
                    item_name = cell.value
                    
            if cell.fill and cell.fill.start_color and cell.fill.start_color.index:
                color_index = str(cell.fill.start_color.index)
                theme = str(cell.fill.start_color.theme)
                if color_index == '8' or theme == '8':
                    is_highlighted = True
        
        if is_highlighted:
            count += 1
            if item_name:
                highlighted_items.append(item_name)
    
    if count > 0:
        print(f"Sheet '{sheetname}': {count} produtos destacados.")
        for item in highlighted_items:
            print(f"  - {item}")
        total_highlighted += count

print(f"\nTotal: {total_highlighted} produtos destacados em todo o Excel.")
