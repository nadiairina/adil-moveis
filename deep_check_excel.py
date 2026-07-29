import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('../Excel de produtos - Site-finalporagora.xlsx', data_only=True)
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ---")
    
    # Just print the first 10 rows to see if there's an extra column
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i < 10:
            print(row)
        
    # Let's also check the exact fill types of the first 20 cells
    print("Checking fill types:")
    for i, row in enumerate(sheet.iter_rows()):
        if i < 20:
            for cell in row:
                if cell.fill and cell.fill.start_color and cell.fill.start_color.index:
                    print(f"Row {cell.row}, Col {cell.column} ({cell.value}): Fill Type = {cell.fill.fill_type}, Color = {cell.fill.start_color.index}, Theme = {cell.fill.start_color.theme}")

