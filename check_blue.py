import openpyxl
import glob

excel_files = glob.glob('../*.xlsx')
for f in excel_files:
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        print(f"--- Checking {f} ---")
        blue_found = 0
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            for row in sheet.iter_rows():
                for cell in row:
                    # check fill color
                    fill = cell.fill
                    font = cell.font
                    if fill and fill.start_color and fill.start_color.index != '00000000' and str(fill.start_color.index) != 'None':
                        # print(f"Fill color: {fill.start_color.index} at {cell.coordinate}")
                        if '0000FF' in str(fill.start_color.index) or 'FF00B0F0' in str(fill.start_color.index) or 'FF0070C0' in str(fill.start_color.index) or 'blue' in str(fill.start_color.index).lower() or '4' in str(fill.start_color.index):
                            blue_found += 1
                    if font and font.color and font.color.index != '00000000' and str(font.color.index) != 'None':
                        if '0000FF' in str(font.color.index) or 'FF00B0F0' in str(font.color.index):
                            blue_found += 1
        if blue_found > 0:
            print(f"Found {blue_found} cells with potential blue highlighting in {f}!")
        else:
            print(f"No blue highlighting found in {f}.")
    except Exception as e:
        print(f"Error reading {f}: {e}")
