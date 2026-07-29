import openpyxl

file_path = '../Excel de produtos - Site-finalporagora.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        # We assume if the first cell of the row has a fill, the row is highlighted
        # Or we can check any cell in the row
        colored_cells = []
        for cell in row:
            if cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000':
                colored_cells.append(cell)
        
        if colored_cells:
            color_index = colored_cells[0].fill.start_color.index
            row_values = [str(c.value) if c.value is not None else '' for c in row]
            # Ignore empty rows
            if any(row_values):
                print(f"Row {row[0].row} [Color: {color_index}]: {row_values}")

