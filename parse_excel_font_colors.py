import openpyxl

file_path = '../Excel de produtos - Site-finalporagora.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        colored_cells = []
        for cell in row:
            if cell.font and cell.font.color and cell.font.color.type == 'rgb' and cell.font.color.rgb != 'FF000000':
                colored_cells.append(cell)
            elif cell.fill and getattr(cell.fill, 'start_color', None) and cell.fill.start_color.index not in ['00000000', 0, 'FFFFFFFF']:
                colored_cells.append(cell)
        
        if colored_cells:
            row_values = [str(c.value) if c.value is not None else '' for c in row]
            if any(row_values):
                print(f"Sheet {sheet_name} - Row {row[0].row} - Values: {row_values}")

